"""Tally Trial Balance (XLSX) → canonical ledger.

Tally lets users export a Trial Balance as XLSX. The exported file is a
flat 3-column sheet:

  Row 1-9  Header (company name, period, column headers)
  Row 10+  ledger_name | debit_amount | credit_amount
  Last row "Grand Total"

This connector parses that, builds the chart of accounts in the
canonical `accounts` table, and posts an opening-balance entry per
ledger that has a non-zero side. After running, the canonical ledger
holds the customer's full balance sheet at the period end — which is
the SHEET TRUTH that bank-statement reconstruction can't see.

Limits of this format:
  * No movement detail (only closing balances) — Day Book is needed for that.
  * No group path per ledger — we infer category from the name. Tally's
    full Trial Balance with Group dropdown enabled does include groups
    in some export modes; we handle both gracefully.

The "smart" classification heuristic for ambiguous names (company names
like "Wellspring Healthcare Private Limited") uses the debit/credit side:
  * Debit balance + company-name shape → receivables (likely customer)
  * Credit balance + company-name shape → payables (likely vendor)
This isn't perfect but matches the ~99% case in real Tally exports
(intercompany loans are the edge case — those should have group paths).
"""
from __future__ import annotations

import logging
import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from common.models import Document, SourceSystem
from services.canonical import accounts as accounts_svc
from services.canonical import ledger as ledger_svc
from services.connectors.base import BaseConnector, HealthResult, SyncResult
from services.connectors.registry import register

logger = logging.getLogger(__name__)


# Indian fiscal-year suffixes Tally uses in headers: "1-Apr-25 to 31-Mar-26"
_PERIOD_RE = re.compile(
    r"(\d{1,2})[-\s]([A-Za-z]{3})[-\s](\d{2,4})\s+to\s+(\d{1,2})[-\s]([A-Za-z]{3})[-\s](\d{2,4})",
    re.I,
)

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


# Heuristic: looks like an Indian/global company name.
# These get the debit/credit smart-classification when the name regex
# can't find a category.
_COMPANY_SUFFIX = re.compile(
    r"\b(pvt\.?\s*ltd\.?|private\s+limited|limited|ltd\.?|llp|inc\.?|corp\.?|"
    r"corporation|company|co\.?|gmbh|s\.a\.|s\.l\.|holdings?|enterprises?|"
    r"associates?|industries?|services?)\b",
    re.I,
)


def _parse_amount(val: Any) -> Decimal:
    """Tally writes numbers either as numerics or as strings with commas.
    Empty / None → Decimal('0'). Bad input → Decimal('0') (we log it).
    """
    if val is None or val == "":
        return Decimal("0")
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", "").replace("₹", "")
    if s.startswith("(") and s.endswith(")"):  # negative in parens
        s = "-" + s[1:-1]
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        logger.warning("Could not parse amount %r as Decimal", val)
        return Decimal("0")


def _parse_period(text: str) -> tuple[Optional[date], Optional[date]]:
    """Parse a Tally period header like '1-Apr-25 to 31-Mar-26'."""
    if not text:
        return None, None
    m = _PERIOD_RE.search(str(text))
    if not m:
        return None, None
    try:
        d1 = date(_expand_year(int(m.group(3))), _MONTH_MAP[m.group(2).lower()[:3]], int(m.group(1)))
        d2 = date(_expand_year(int(m.group(6))), _MONTH_MAP[m.group(5).lower()[:3]], int(m.group(4)))
        return d1, d2
    except (KeyError, ValueError):
        logger.warning("Could not parse period text %r", text)
        return None, None


def _expand_year(y: int) -> int:
    """'25' → 2025; '99' → 1999 (Tally is post-2000 in practice)."""
    if y >= 100:
        return y
    return 2000 + y if y < 80 else 1900 + y


def _looks_like_company_name(name: str) -> bool:
    if not name:
        return False
    if _COMPANY_SUFFIX.search(name):
        return True
    # Multi-word, mostly proper-cased words with no expense/income keywords.
    parts = name.strip().split()
    if len(parts) >= 3 and sum(1 for p in parts if p[:1].isupper()) >= len(parts) // 2:
        # And nothing screams "expense" or "income"
        EXPENSE_HINTS = {"rent", "salary", "expense", "fee", "fees", "charge", "charges", "tax", "duty"}
        if not any(p.lower() in EXPENSE_HINTS for p in parts):
            return True
    return False


def _smart_classify(name: str, debit: Decimal, credit: Decimal) -> str:
    """Classify a ledger when no group path is available. Falls back to
    the debit/credit-side hint for company-name-shaped accounts.
    """
    cat, _ = accounts_svc.classify(name)
    if cat != "suspense":
        return cat
    # Company-shaped names with a balance → debtor/creditor inference
    if _looks_like_company_name(name):
        if debit > credit:
            return "receivables"
        if credit > debit:
            return "payables"
    return "suspense"


@register
class TallyTrialBalanceConnector(BaseConnector):
    """Connector that ingests a Tally Trial Balance XLSX upload."""

    system_type = "tally_trial_balance"
    display_name = "Tally — Trial Balance (XLSX upload)"
    category = "ledger"
    supports_file_upload = True

    def health_check(self) -> HealthResult:
        # File-upload-only connector; nothing to ping. Always green.
        return HealthResult(ok=True, detail="upload-only connector")

    def ingest_file(
        self,
        file_path: str,
        mime_type: Optional[str] = None,
        original_filename: Optional[str] = None,
        document_id: Optional[uuid.UUID] = None,
    ) -> SyncResult:
        try:
            import openpyxl
        except ImportError:
            return SyncResult(errors=["openpyxl not installed"])

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:
            return SyncResult(errors=[f"could not open XLSX: {exc}"])

        ws = wb.active
        if ws is None:
            return SyncResult(errors=["workbook has no active sheet"])

        # ---- Parse header block ---------------------------------------------
        title_row = ws.cell(row=1, column=1).value or ""
        period_text = ws.cell(row=5, column=1).value or ws.cell(row=4, column=2).value or ""
        period_start, period_end = _parse_period(period_text)
        if period_end is None:
            # Fallback: use the file's modified date or today
            period_end = date.today()
        as_of = period_end

        result = SyncResult()
        result.detail["company_name"] = str(title_row).strip() if title_row else None
        result.detail["period_text"] = str(period_text).strip() if period_text else None
        result.detail["as_of"] = as_of.isoformat()

        # ---- Find data start: first row where col A is non-empty AND
        # col B or C is numeric. Tally typically starts at row 10.
        data_start = None
        for r in range(1, min(ws.max_row + 1, 50)):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            if a and not str(a).lower().startswith(("particulars", "grand total")):
                if isinstance(b, (int, float)) or isinstance(c, (int, float)):
                    data_start = r
                    break
        if data_start is None:
            result.errors.append("could not find start of data rows in TB sheet")
            return result

        # ---- Iterate rows ---------------------------------------------------
        total_dr = Decimal("0")
        total_cr = Decimal("0")
        rows_processed = 0
        suspense_rows: list[str] = []

        for r in range(data_start, ws.max_row + 1):
            name_cell = ws.cell(row=r, column=1).value
            if name_cell is None:
                continue
            name = str(name_cell).strip()
            if not name:
                continue
            # Stop at Grand Total
            if name.lower().startswith("grand total"):
                break

            dr = _parse_amount(ws.cell(row=r, column=2).value)
            cr = _parse_amount(ws.cell(row=r, column=3).value)
            if dr == 0 and cr == 0:
                continue

            # Classify
            category = _smart_classify(name, dr, cr)
            if category == "suspense":
                suspense_rows.append(name)

            # Find or create the account
            account = accounts_svc.find_or_create(
                self.db,
                org_id=self.org_id,
                entity_id=self.entity_id,
                name=name,
                source_system_id=self.source_system.id,
                source_native_id=f"tally_tb:{name}",
                hinted_category=category,
                commit=False,
            )
            result.accounts_upserted += 1

            # Post an opening-balance ledger entry (single-leg, suspense mirror)
            try:
                ledger_svc.post_opening_balance(
                    self.db,
                    org_id=self.org_id,
                    entity_id=self.entity_id,
                    account_id=account.id,
                    as_of=as_of,
                    debit_inr=dr,
                    credit_inr=cr,
                    source_system_id=self.source_system.id,
                    source_document_id=document_id,
                    commit=False,
                )
                result.transactions_written += 1
                result.ledger_entries_written += 2  # ledger leg + suspense mirror
            except Exception as exc:
                result.errors.append(f"row {r} '{name}': {exc}")

            total_dr += dr
            total_cr += cr
            rows_processed += 1

        # Commit everything in a single transaction
        try:
            self.db.commit()
        except Exception as exc:
            self.db.rollback()
            result.errors.append(f"commit failed: {exc}")
            return result

        result.detail["rows_processed"] = rows_processed
        result.detail["total_debit"] = str(total_dr)
        result.detail["total_credit"] = str(total_cr)
        result.detail["delta"] = str(total_dr - total_cr)
        result.detail["suspense_count"] = len(suspense_rows)
        result.detail["suspense_samples"] = suspense_rows[:10]

        # If the file's debits and credits don't match, the source TB was
        # already unbalanced — flag it but don't fail.
        if abs(total_dr - total_cr) > Decimal("1.00"):
            result.errors.append(
                f"source TB unbalanced: Dr {total_dr} vs Cr {total_cr} "
                f"(Δ {total_dr - total_cr})"
            )

        # Cursor: remember when we last ingested + which period
        result.cursor = {
            "last_period_end": as_of.isoformat(),
            "last_ingested_at": datetime.utcnow().isoformat(),
            "last_total_inr": str(max(total_dr, total_cr)),
        }
        return result


# ---------------------------------------------------------------------------
# Helper: convenience function used by the upload route.
#
# Lets routes import a single function rather than build a ConnectorContext
# themselves. Internally builds the SourceSystem row + context.
# ---------------------------------------------------------------------------


def ingest_trial_balance_xlsx(
    db,
    org_id: uuid.UUID,
    file_path: str,
    document_id: Optional[uuid.UUID] = None,
    entity_id: Optional[uuid.UUID] = None,
    display_name: Optional[str] = None,
    original_filename: Optional[str] = None,
) -> SyncResult:
    """Top-level helper. Finds or creates a SourceSystem row for the
    'tally_trial_balance' type, then runs the connector.

    Returns the SyncResult so the API can show counts + suspense
    examples to the user.
    """
    from sqlalchemy import select

    from services.canonical import entities as entities_svc

    entity = entities_svc.resolve_entity(db, org_id, entity_id)
    src = db.execute(
        select(SourceSystem).where(
            SourceSystem.org_id == org_id,
            SourceSystem.entity_id == entity.id,
            SourceSystem.system_type == "tally_trial_balance",
        ).limit(1)
    ).scalar_one_or_none()

    if src is None:
        src = SourceSystem(
            org_id=org_id,
            entity_id=entity.id,
            system_type="tally_trial_balance",
            display_name=display_name or "Tally Trial Balance (manual upload)",
            config_json={},
            cursor_json={},
            is_enabled=True,
        )
        db.add(src)
        db.commit()
        db.refresh(src)

    from services.connectors.base import ConnectorContext

    ctx = ConnectorContext(db=db, org_id=org_id, entity_id=entity.id, source_system=src)
    connector = TallyTrialBalanceConnector(ctx)
    result = connector.ingest_file(
        file_path,
        original_filename=original_filename,
        document_id=document_id,
    )

    # Mirror result into source_systems row
    src.last_sync_at = datetime.utcnow()
    src.last_sync_status = result.status
    src.last_sync_error = "; ".join(result.errors)[:500] if result.errors else None
    if result.cursor:
        merged = dict(src.cursor_json or {})
        merged.update(result.cursor)
        src.cursor_json = merged
    db.commit()
    return result
