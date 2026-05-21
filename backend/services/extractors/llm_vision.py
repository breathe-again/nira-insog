"""Anthropic-vision-backed document extractor.

Replaces the Tesseract+LLM path with a single API call: send the PDF/image
to Claude and get back a JSON payload that matches the shape our
`extracted_json` parser already understands.

Why a single LLM call:
- One round-trip instead of OCR → text → LLM.
- Better accuracy on Indian invoices/receipts where layouts are messy.
- No native dependencies (no Tesseract, no Poppler) → smaller image.

Cost (May 2026):
- ~₹0.20–0.60 per single-page invoice/receipt.
- ~₹2–5 for a 10-page bank statement PDF.

Disabled when ANTHROPIC_API_KEY is unset → caller should fall back to stub.
"""

from __future__ import annotations

import base64
import json
import logging
import os
import re
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# Model — pinned to the latest Sonnet at time of writing.
DEFAULT_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-6")

# Hard cap on response size. Long bank statements with 100+ transactions
# can hit ~8k tokens, so we generously allow 16k to avoid mid-JSON truncation.
MAX_TOKENS = 16000


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def is_enabled() -> bool:
    """True if the extractor is configured. Cheap — just an env check."""
    return bool(os.environ.get("ANTHROPIC_API_KEY"))


class ExtractorError(RuntimeError):
    """Raised when the extractor can't produce usable output."""


def extract(
    file_path: Path,
    *,
    file_type: str,                 # "pdf" | "image" | "html" | "xlsx"
    document_type_hint: str = "unknown",
) -> dict:
    """Send a single document to Claude and return the extracted JSON.

    The JSON shape matches what `services/parsers/extracted_json.py`
    expects — see that module's docstring for the contract.

    Supported file_type values:
      - "pdf"   → sent as a base64 PDF document block (Claude vision).
      - "image" → sent as a base64 image block (Claude vision).
      - "html"  → decoded to text and stripped of tags; sent as a text block.
      - "xlsx"  → workbook flattened to a CSV-like text dump; sent as a text block.

    Raises ExtractorError if the API key isn't set, the API call fails, or
    the model returns something we can't parse as JSON.
    """
    if not is_enabled():
        raise ExtractorError("ANTHROPIC_API_KEY is not set")

    if not file_path.exists():
        raise ExtractorError(f"file not on disk: {file_path}")

    if file_type not in ("pdf", "image", "html", "xlsx"):
        raise ExtractorError(f"unsupported file_type={file_type!r}")

    try:
        # Lazy import so the rest of the worker still loads when anthropic
        # isn't installed (e.g. local dev without LLM features).
        import anthropic  # type: ignore
    except ImportError as e:  # pragma: no cover
        raise ExtractorError(
            "anthropic package not installed. Add to requirements.txt."
        ) from e

    client = anthropic.Anthropic()  # reads ANTHROPIC_API_KEY from env

    content = _build_content_blocks(file_path, file_type)
    prompt = _build_prompt(document_type_hint)

    logger.info(
        "llm_vision: extracting %s (%s, %d bytes)",
        file_path.name,
        file_type,
        file_path.stat().st_size,
    )

    try:
        response = client.messages.create(
            model=DEFAULT_MODEL,
            max_tokens=MAX_TOKENS,
            system=_SYSTEM_PROMPT,
            messages=[
                {
                    "role": "user",
                    "content": [*content, {"type": "text", "text": prompt}],
                }
            ],
        )
    except Exception as e:  # noqa: BLE001
        raise ExtractorError(f"Anthropic API call failed: {e}") from e

    text = _flatten_response(response)
    payload = _parse_json(text)

    # Attach provenance so downstream code knows where this came from.
    payload.setdefault("_meta", {})
    payload["_meta"].update({
        "extractor": "anthropic_vision",
        "model": DEFAULT_MODEL,
        "tokens_in": getattr(response.usage, "input_tokens", None),
        "tokens_out": getattr(response.usage, "output_tokens", None),
    })

    return payload


# ---------------------------------------------------------------------------
# Content + prompt construction
# ---------------------------------------------------------------------------


def _build_content_blocks(file_path: Path, file_type: str) -> list[dict]:
    """Build the list of content blocks for the API message."""
    if file_type == "html":
        text = _html_to_text(file_path)
        return [
            {
                "type": "text",
                "text": (
                    "Document source: HTML (e.g. a bank's web export). "
                    "The visible text content follows between <<<HTML>>> markers; "
                    "navigation chrome and scripts have been stripped.\n\n"
                    "<<<HTML>>>\n" + text + "\n<<<END>>>"
                ),
            }
        ]

    if file_type == "xlsx":
        text = _xlsx_to_text(file_path)
        return [
            {
                "type": "text",
                "text": (
                    "Document source: Excel workbook (.xlsx/.xls). "
                    "Each sheet is dumped below as tab-separated rows.\n\n"
                    "<<<WORKBOOK>>>\n" + text + "\n<<<END>>>"
                ),
            }
        ]

    data = file_path.read_bytes()
    b64 = base64.standard_b64encode(data).decode("ascii")

    if file_type == "pdf":
        return [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": b64,
                },
            }
        ]

    # file_type == "image"
    suffix = file_path.suffix.lower()
    media_type = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }.get(suffix, "image/jpeg")
    return [
        {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_type,
                "data": b64,
            },
        }
    ]


# ---------------------------------------------------------------------------
# Text-extraction helpers for non-vision file types
# ---------------------------------------------------------------------------


# Cap how much text we send to the LLM. Bank HTML exports can be large; Claude
# Sonnet's context is huge, but we want to keep cost predictable.
_MAX_TEXT_CHARS = 200_000


def _html_to_text(file_path: Path) -> str:
    """Read an HTML file and return its visible text.

    No external HTML parser dependency — uses stdlib `html.parser`. We strip
    <script> / <style> contents and collapse whitespace. Tables (very common
    in bank HTML exports) are flattened with tab separators between cells and
    newlines between rows, which preserves enough structure for the LLM to
    read out transactions.
    """
    raw = file_path.read_bytes()
    # Most banks emit UTF-8 or windows-1252. Try both.
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    from html.parser import HTMLParser

    SKIP_TAGS = {"script", "style", "noscript", "head", "meta", "link"}
    BLOCK_TAGS = {
        "p", "div", "br", "li", "h1", "h2", "h3", "h4", "h5", "h6",
        "section", "article", "header", "footer", "aside", "nav",
    }

    class _Collector(HTMLParser):
        def __init__(self) -> None:
            super().__init__(convert_charrefs=True)
            self.out: list[str] = []
            self._skip_depth = 0

        def handle_starttag(self, tag: str, attrs):  # noqa: ANN001
            if tag in SKIP_TAGS:
                self._skip_depth += 1
                return
            if tag == "tr":
                self.out.append("\n")
            elif tag in ("td", "th"):
                self.out.append("\t")
            elif tag == "table":
                self.out.append("\n")
            elif tag in BLOCK_TAGS:
                self.out.append("\n")

        def handle_endtag(self, tag: str):
            if tag in SKIP_TAGS and self._skip_depth > 0:
                self._skip_depth -= 1
                return
            if tag in ("table", "tr"):
                self.out.append("\n")

        def handle_data(self, data: str):
            if self._skip_depth > 0:
                return
            if data.strip():
                self.out.append(data)

    collector = _Collector()
    try:
        collector.feed(text)
    except Exception:  # noqa: BLE001
        # Malformed HTML — fall through with whatever we got so far.
        pass

    flat = "".join(collector.out)
    # Collapse runs of blank lines and internal whitespace within lines.
    lines = []
    for line in flat.splitlines():
        compact = " ".join(line.split())
        if compact:
            lines.append(compact)
    out = "\n".join(lines)

    if len(out) > _MAX_TEXT_CHARS:
        out = out[:_MAX_TEXT_CHARS] + "\n... [truncated]"
    return out


def _xlsx_to_text(file_path: Path) -> str:
    """Read an .xlsx/.xls workbook and return a tab-separated dump.

    Uses openpyxl (read-only) — pulls cell values across every sheet, one row
    per line, tabs between cells. .xls (legacy) isn't supported by openpyxl;
    we surface a clear error so the caller can fall back to the stub.
    """
    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        raise ExtractorError(
            "legacy .xls files not supported — please re-save as .xlsx"
        )

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:  # pragma: no cover
        raise ExtractorError(
            "openpyxl not installed — required for .xlsx extraction"
        ) from e

    try:
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ExtractorError(f"could not open workbook: {e}") from e

    chunks: list[str] = []
    total_chars = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(f"\n## Sheet: {sheet_name}\n")
        for row in ws.iter_rows(values_only=True):
            # Skip rows that are entirely empty.
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue
            cells = ["" if c is None else str(c) for c in row]
            line = "\t".join(cells) + "\n"
            chunks.append(line)
            total_chars += len(line)
            if total_chars > _MAX_TEXT_CHARS:
                chunks.append("... [truncated]\n")
                wb.close()
                return "".join(chunks)
    wb.close()
    return "".join(chunks)


_SYSTEM_PROMPT = """\
You are an extraction engine for an Indian SMB accounting platform.
You read invoices, receipts, and bank statements and return structured JSON.

Rules:
- Output ONLY a single JSON object. No prose, no markdown fences, no explanation.
- Use ISO dates (YYYY-MM-DD). Amounts are plain numbers (no currency symbols, no commas).
- For Indian GSTINs, normalize to uppercase. Omit the field if not visible.
- If a field is not present, omit it (do not invent values).
"""


def _build_prompt(document_type_hint: str) -> str:
    """User message text. The hint is best-effort — model can override."""
    hint = ""
    if document_type_hint and document_type_hint != "unknown":
        hint = f"\nFile-name hint: this is likely a {document_type_hint}.\n"

    return f"""\
Extract the contents of this document into a single JSON object.

Choose ONE of these document_type values based on what you actually see:
  - "purchase_invoice"  — invoice issued TO us by a vendor (a bill we owe)
  - "sales_invoice"     — invoice WE issued to a customer
  - "receipt"           — a paid receipt / cash memo
  - "bank_statement"    — a list of bank transactions
{hint}
For "purchase_invoice" / "sales_invoice" emit:
  {{
    "document_type": "purchase_invoice",
    "invoice_number": "...",
    "vendor":  {{"name": "...", "gstin": "..."}}   // for purchase
    "client":  {{"name": "...", "gstin": "..."}}   // for sales
    "issue_date": "YYYY-MM-DD",
    "due_date":   "YYYY-MM-DD",
    "currency":   "INR",
    "subtotal":   <number>,
    "tax":        <number>,
    "total":      <number>,
    "line_items": [{{"description": "...", "qty": 1, "unit_price": 100, "amount": 100}}, ...]
  }}

For "receipt" emit:
  {{
    "document_type": "receipt",
    "vendor": {{"name": "..."}},
    "date":   "YYYY-MM-DD",
    "amount": <number>,
    "tax":    <number>,
    "category": "meals" | "travel" | "office" | "marketing" | "utilities" | "other",
    "payment_mode": "cash" | "card" | "upi" | "bank_transfer" | "unknown"
  }}

For "bank_statement" emit:
  {{
    "document_type": "bank_statement",
    "account_holder": "...",
    "account_number_last4": "1234",
    "currency": "INR",
    "period_start": "YYYY-MM-DD",
    "period_end":   "YYYY-MM-DD",
    "transactions": [
      {{"date": "YYYY-MM-DD", "description": "...", "amount": <number>, "direction": "debit" | "credit", "balance": <number>}}
    ]
  }}

Return ONLY the JSON object. Begin your response with `{{` and end with `}}`.
"""


# ---------------------------------------------------------------------------
# Response parsing
# ---------------------------------------------------------------------------


def _flatten_response(response) -> str:  # type: ignore[no-untyped-def]
    """Extract the concatenated text content from a Messages API response."""
    parts: list[str] = []
    for block in response.content:
        # The SDK returns content blocks; for text blocks, `.text` holds the value.
        text = getattr(block, "text", None)
        if text:
            parts.append(text)
    return "".join(parts)


_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(\{.*?\})\s*```", re.DOTALL)


def _parse_json(text: str) -> dict:
    """Best-effort JSON parse — tolerates stray fences / leading whitespace."""
    if not text:
        raise ExtractorError("empty response from model")

    candidate = text.strip()

    # If the model wrapped it in ```json fences, unwrap.
    m = _JSON_FENCE_RE.search(candidate)
    if m:
        candidate = m.group(1)

    # Last-ditch: take from the first `{` to the last `}`.
    first = candidate.find("{")
    last = candidate.rfind("}")
    if first >= 0 and last > first:
        candidate = candidate[first : last + 1]

    try:
        parsed = json.loads(candidate)
    except json.JSONDecodeError as e:
        # LLM JSON often has small flaws (stray comma, unescaped quote in a
        # transaction narration, etc.). Try to repair before giving up.
        try:
            from json_repair import repair_json  # type: ignore

            repaired = repair_json(candidate, return_objects=True)
            if isinstance(repaired, dict):
                logger.warning(
                    "llm_vision: stdlib JSON parse failed (%s); json_repair recovered the payload",
                    e,
                )
                return repaired
        except ImportError:
            pass
        except Exception:  # noqa: BLE001
            pass

        raise ExtractorError(
            f"model returned non-JSON output: {e}; first 200 chars: {text[:200]!r}"
        ) from e

    if not isinstance(parsed, dict):
        raise ExtractorError(f"expected JSON object, got {type(parsed).__name__}")

    return parsed


# ---------------------------------------------------------------------------
# Convenience helpers used by the Celery task
# ---------------------------------------------------------------------------


def extract_safely(
    file_path: Path,
    *,
    file_type: str,
    document_type_hint: str = "unknown",
) -> Optional[dict]:
    """Wrapper that returns None instead of raising — convenient for callers
    that want to fall back to the stub on any failure."""
    try:
        return extract(file_path, file_type=file_type, document_type_hint=document_type_hint)
    except ExtractorError as e:
        logger.warning("llm_vision extract skipped (%s): %s", file_path.name, e)
        return None
    except Exception as e:  # noqa: BLE001 — never crash the worker on extractor bugs
        logger.exception("llm_vision unexpected error (%s)", file_path.name)
        _ = e
        return None
