"""Unit tests for the Tally Trial Balance connector.

The DB-touching parts are covered by integration tests; here we
exercise the pure-Python helpers (amount/period parsing, smart
classification) against the actual Quantta TrialBal.xlsx file when
it's available.
"""
from __future__ import annotations

import os
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from services.connectors.tally_trial_balance import (
    _looks_like_company_name,
    _parse_amount,
    _parse_period,
    _smart_classify,
)


@pytest.mark.parametrize(
    "raw,expected",
    [
        (None, Decimal("0")),
        ("", Decimal("0")),
        (1234.56, Decimal("1234.56")),
        ("1,234.56", Decimal("1234.56")),
        ("₹ 1,234.56", Decimal("1234.56")),
        ("(500.00)", Decimal("-500.00")),
        ("abc", Decimal("0")),
        (281373717.86, Decimal("281373717.86")),
    ],
)
def test_parse_amount(raw, expected):
    assert _parse_amount(raw) == expected


@pytest.mark.parametrize(
    "text,expected",
    [
        ("1-Apr-25 to 31-Mar-26", (date(2025, 4, 1), date(2026, 3, 31))),
        ("1-Apr-2025 to 31-Mar-2026", (date(2025, 4, 1), date(2026, 3, 31))),
        ("Q1 - Apr to Jun", (None, None)),  # not the expected format
        ("", (None, None)),
        (None, (None, None)),
    ],
)
def test_parse_period(text, expected):
    assert _parse_period(text) == expected


@pytest.mark.parametrize(
    "name,expected",
    [
        ("Amazon Web Services India Private Limited", True),
        ("Wellspring Healthcare Private Limited", True),
        ("Lichee Construction Pvt Ltd", True),
        ("Quantta Analytics Pvt. Ltd.", True),
        ("Apple Inc.", True),
        ("Bawri Logistics LLP", True),
        ("Vinay Gopinath", False),  # individual name
        ("Salary", False),
        ("Bank Charges", False),
        ("HDFC Bank", False),  # has bank category
        ("Office Rent", False),
    ],
)
def test_looks_like_company(name, expected):
    assert _looks_like_company_name(name) == expected


def test_smart_classify_debit_company_is_receivable():
    # Company-shaped name with debit balance → receivable (customer)
    cat = _smart_classify(
        "Wellspring Healthcare Private Limited",
        debit=Decimal("351235.95"),
        credit=Decimal("0"),
    )
    assert cat == "receivables"


def test_smart_classify_credit_company_is_payable():
    # Company-shaped name with credit balance → payable (vendor)
    cat = _smart_classify(
        "Amazon Web Services India Private Limited",
        debit=Decimal("0"),
        credit=Decimal("139245.79"),
    )
    assert cat == "payables"


def test_smart_classify_known_account_overrides_company_heuristic():
    # "HDFC Bank ..." matches the bank rule — must classify as bank
    # regardless of debit/credit side.
    cat = _smart_classify(
        "HDFC Bank Quantta Analytics Private Limited",
        debit=Decimal("100"), credit=Decimal("0"),
    )
    assert cat == "bank"


def test_smart_classify_individual_name_falls_through():
    # Individual name with no expense/income hint → stays suspense
    cat = _smart_classify("Vinay Gopinath", debit=Decimal("0"), credit=Decimal("131988"))
    # We treat individuals as suspense; the inverse case (people who
    # ARE customers/vendors) gets resolved by a user-side merge.
    assert cat == "suspense"


# -----------------------------------------------------------------------------
# Integration smoke test against the real Quantta TrialBal.xlsx if present.
# Skips automatically when the file isn't on disk (e.g. in CI).
# -----------------------------------------------------------------------------

_QUANTTA_FILES = [
    Path("/sessions/determined-sweet-faraday/mnt/uploads/TrialBal.xlsx"),
    Path("/uploads/TrialBal.xlsx"),
    Path("./TrialBal.xlsx"),
]


def _find_quantta_file() -> Path | None:
    for p in _QUANTTA_FILES:
        if p.exists():
            return p
    return None


@pytest.mark.skipif(_find_quantta_file() is None, reason="TrialBal.xlsx not on disk")
def test_quantta_xlsx_parses_balanced():
    """End-to-end sanity: every row of the real Trial Balance parses and
    the file's grand total matches the row sum (₹28.13 Cr)."""
    import openpyxl

    path = _find_quantta_file()
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb.active

    total_dr = Decimal("0")
    total_cr = Decimal("0")
    rows = 0
    grand_dr = Decimal("0")
    grand_cr = Decimal("0")
    for r in range(10, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        name_s = str(name).strip()
        if name_s.lower().startswith("grand total"):
            grand_dr = _parse_amount(ws.cell(row=r, column=2).value)
            grand_cr = _parse_amount(ws.cell(row=r, column=3).value)
            break
        dr = _parse_amount(ws.cell(row=r, column=2).value)
        cr = _parse_amount(ws.cell(row=r, column=3).value)
        if dr == 0 and cr == 0:
            continue
        total_dr += dr
        total_cr += cr
        rows += 1

    assert rows > 50, f"expected 100+ ledger rows, got {rows}"
    assert grand_dr == total_dr, (
        f"row sum debit {total_dr} != Grand Total debit {grand_dr}"
    )
    assert grand_cr == total_cr, (
        f"row sum credit {total_cr} != Grand Total credit {grand_cr}"
    )
    # Trial Balance must self-balance
    assert grand_dr == grand_cr, (
        f"TB unbalanced: Dr {grand_dr} vs Cr {grand_cr}"
    )
